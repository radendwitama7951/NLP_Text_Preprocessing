"""
Program Natural Language Preproccessing
by Raden Dwitama Baliano
23/2/2022

"""
# Library
import numpy as np;
import matplotlib.pyplot as plt;
from os import path, makedirs;
from openpyxl import load_workbook, Workbook;
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter;
from openpyxl.styles import Font;
from nltk import FreqDist, pprint;
from nltk.tokenize import regexp_tokenize;
from nltk.corpus import stopwords;
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory;
from copy import copy;
from itertools import chain;
from inspect import isclass;
from shutil import copyfile;

# Global
FILEPATH_RAW = "./DATA RESULT/DatasetText.xlsx";
FILEPATH_RESULT = "./DATA RESULT/Hasil.xlsx";

###################### UTILS #####################
"""
Membuat file excel baru
agar file mentah tidak rusak saat percobaan
"""
def copy_excel (source_sheet, n_row, target_filename):        
    # Definisikan workbook baru
    workbook_baru = Workbook();
    sheet_baru = workbook_baru.active;
    sheet_baru.title = source_sheet.title;


    # Hitung jumlah baris dan kolom sheet asal
    count_source_row = n_row;
    count_source_col = source_sheet.max_column;

    # Copy cell asal ke cell sheet baru
    for i in range(1, count_source_row + 1):
        for j in range(1, count_source_col + 1):
            # Ambil value cellnya
            source_cell = source_sheet.cell(row = i, column = j);

            # Copy value source cell ke cell baru
            sheet_baru.cell(row = i, column = j).value = source_cell.value;
    workbook_baru.save(filename = target_filename);
    return workbook_baru;

"""
Mendapatkan koordinat dalam bentuk index
"""
def get_idx_coordinate (cell):
    # Ambil coordinate dalam forma "A1","A2", dsb.
    cell_idx_coordinate = coordinate_from_string(cell);

    # Ubah format alfabet (A,B,C, dsb.) dalam coordinate menjadi index angka
    cell_idx_coordinate = [column_index_from_string(cell_idx_coordinate[0]), cell_idx_coordinate[1]];

    # Return tupple index
    return cell_idx_coordinate;
    

"""
Mendapatkan column berdasarkan nama field
"""
def get_column_by_value (sheet, col_to_find):
    for col_index in range(1, sheet.max_column+1):
        if sheet.cell(1, col_index).value == col_to_find:
            return col_index;
        
    print("Column dengan value " + col_to_find + " tidak ditemukan !");
    quit();


"""
Ambil tabel dari sheet workbook
menjadi list []
"""
def get_list_from_table (sheet, cell_1, cell_2):
    # List penampung untuk di return
    table_list = [];

    # Ambil cell pada pojok kiri atas tabel yang hendak di select
    cell_1 = get_idx_coordinate(cell_1);
    cell_2 = get_idx_coordinate(cell_2);

    # List sementara untuk menyimpan column


    # Iterasi setiap cell pada rentang cell_1 dan cell_2 
    for row in sheet.iter_rows(min_row = cell_1[1], max_row = cell_2[1], min_col = cell_1[0]):
        row_list = [];
        for cell in row:
            if not cell.value is None:
                # PUSH setiap column yang berisi value saja
                # ke dalam List row (list sementara)
                row_list.append(cell.value);

        # PUSH list value setiap column Token ke dalam table_list
        # setiap membaca satu baris
        table_list.append(row_list);


    
    return table_list; 

"""
Membuat heading setiap field menjadi bold
"""
def set_heading_to_bold (workbook, target_sheet):
    # Ambil sheet target
    if isinstance(target_sheet, str):
       target_sheet = workbook[target_sheet];

    # Format heading
    heading_style = Font(bold = True);

    # Iterasi Cell pada baris pertama
    for cell in target_sheet["1:1"]:
        cell.font = heading_style;

    # Update Workbook
    workbook.save(filename = FILEPATH_RESULT);

 

 

#########################################################        
    
"""
Modul untuk mendapatkan dan memetakan token dari
worksheet Dataset ke worksheet baru
"""
def get_tokenize (source_workbook, source_sheet, target_sheet, cell_1, cell_2):
    # AMbil sheet 1
    source_sheet = source_workbook[source_sheet];

    # Buat sheet hasil tokenisasi
    target_sheet = source_workbook.create_sheet(title = target_sheet);

    # Mengambil kordinat rentang tabel
    cell_1 = get_idx_coordinate(cell_1);
    cell_2 = get_idx_coordinate(cell_2);

    # Simpan kalimat tertokenisasi
    token_list = [];
   
    # Tokenisasi tiap kalimat pada kolom kalimat dan tulis ke kolom token
    for row in source_sheet.iter_rows(min_row = cell_1[1], min_col = cell_1[0], max_row = cell_2[1], max_col = cell_2[0], values_only = True):
        for cell in row:
            # Rule Tokenisasi untuk menghapus tanda baca
            token = regexp_tokenize(cell, "[\w']+");

            # Simpan hasil tokenisasi ke list token
            token_list.append(token);

            # print(tokenized_text);
            target_sheet.append(token);

    # Buat baris baru untuk header
    target_sheet.insert_rows(1);
    
    # Buat kolom baru untuk index nomor kalimat
    target_sheet.insert_cols(1);


    # Merge Header Token pada spreadsheet
    target_sheet.merge_cells(start_row = 1, start_column = cell_1[0], end_row = 1, end_column = target_sheet.max_column);

    # Beri judul pada header kolom token
    target_sheet.cell(row = 1, column = cell_1[0]).value = "Token"; 
    
    # Buat kolom nomor pada sheet token
    i = 1;
    for nomor in source_sheet.iter_rows(min_row = 1, min_col = 1, max_col = 1, values_only = True):
        target_sheet.cell(row = i, column = 1).value = nomor[0];
        i = i + 1;
         


    # Simpan data pada workbook (.xlsx)
    source_workbook.save(filename = FILEPATH_RESULT);

    # Buat heading setiap kolom menjadi bold
    set_heading_to_bold (source_workbook, target_sheet);

    # Return List Token
    return token_list;


"""
Modul untuk mengubah semua karakter
pada rentang cell_1 dan cell_1
menjadi huruf kecil
"""
def set_table_to_lower_case (source_workbook, source_sheet, cell_1, cell_2):
    # Mengambil sheet asal
    source_sheet = source_workbook[source_sheet];
    
    # Mengambil kordinat rentang tabel
    cell_1 = get_idx_coordinate(cell_1);
    cell_2 = get_idx_coordinate(cell_2);


    # Ambil list berdasarkan cell yang dipilih
    for row in source_sheet.iter_rows(min_row = cell_1[1], min_col = cell_1[0], max_row = cell_2[1], max_col = cell_2[0]):
        for cell in row:
            if not cell.value is None:
                cell.value = str(cell.value).lower();
                # print(cell.value);

    # Simpan data pada workbook (.xlsx)
    source_workbook.save(filename = FILEPATH_RESULT);


"""
Modul untuk mendapatkan list dan membuat sheet kata tanpa stopword
berdasarkan rentang tabel pada worksheet token
"""
def get_stopwords_removal (source_workbook, source_sheet, target_sheet, cell_1, cell_2):
    # Mengambil source sheet
    source_sheet = source_workbook[source_sheet];

    # Ambil nilai pada spreadsheet bagian token sebagai tuple
    token_list = get_list_from_table(source_sheet, cell_1, cell_2);

    # Mengambil kordinat rentang tabel
    cell_1 = get_idx_coordinate(cell_1);
    cell_2 = get_idx_coordinate(cell_2);
 
    # Buat sheet baru
    target_sheet = source_workbook.create_sheet(title = target_sheet);

    # Set Stop word
    stop_words = set(stopwords.words("indonesian"));

    # Simpan token tanpa stopwords
    token_no_stopwords_list = [];

    # Iterasi semua token_list
    for token in token_list:
        # Hapus stopwords Bahasa Indoneisia
        token_no_stopwords = [kata for kata in token if not kata in stop_words];

        # Simpan ke list no stopwords   
        token_no_stopwords_list.append( token_no_stopwords );

        # Tulis ke sheet no stopwords (file excel)
        target_sheet.append(token_no_stopwords);

    # Buat baris baru untuk header
    target_sheet.insert_rows(1);
    
    # Buat kolom baru untuk index nomor kalimat
    target_sheet.insert_cols(1);
 
    # Salin no dari sheet source
    i = 1;
    for nomor in source_sheet.iter_rows(min_row = 1, min_col = 1, max_col = 1, values_only = True):
        target_sheet.cell(row = i, column = 1).value = nomor[0];
        i = i + 1;

    # Merge header kolom token
    target_sheet.merge_cells(start_row = 1, start_column = cell_1[0], end_row = 1, end_column = target_sheet.max_column);

    # Tambah judul header
    target_sheet.cell(row = 1, column = cell_1[0]).value = "Token No Stopwords";  




      
    # Update Workbook hasil (file xlsx)
    source_workbook.save(filename = FILEPATH_RESULT);

    # Buat heading setiap kolom menjadi bold
    set_heading_to_bold (source_workbook, target_sheet);

    # Return list token tanpa stopwords
    return token_no_stopwords_list 

        

"""
Modul untuk mendapatkan dan mensteming token pada sheet
yang sudah dihilangkan stopwordsnya
"""
def get_stemming (source_workbook, source_sheet, target_sheet, cell_1, cell_2):
    # Mengambil source sheet
    source_sheet = source_workbook[source_sheet];

    # Ambil nilai pada spreadsheet bagian token sebagai tuple
    token_list = get_list_from_table(source_sheet, cell_1, cell_2);

    # Mengambil kordinat rentang tabel
    cell_1 = get_idx_coordinate(cell_1);
    cell_2 = get_idx_coordinate(cell_2);
 
    # Buat sheet baru
    target_sheet = source_workbook.create_sheet(title = target_sheet);

    # Buat alat Stemmer
    factory = StemmerFactory();
    stemmer = factory.create_stemmer();

    # Penampung hasil stem
    token_stemmed_list = [];

    # Iterasi semua token_list
    for token in token_list:
        # Hapus stopwords Bahasa Indoneisia
        token_stemmed = list(map(lambda kata: stemmer.stem(kata), token));

        # Masukan stemmed token ke dalam list
        token_stemmed_list.append(token_stemmed);

        # Masukan hasil
        target_sheet.append(token_stemmed);        



    # Buat baris baru untuk header
    target_sheet.insert_rows(1);
    
    # Buat kolom baru untuk index nomor kalimat
    target_sheet.insert_cols(1);
 
    # Salin no dari sheet source
    i = 1;
    for nomor in source_sheet.iter_rows(min_row = 1, min_col = 1, max_col = 1, values_only = True):
        target_sheet.cell(row = i, column = 1).value = nomor[0];
        i = i + 1;

    # Merge header kolom token
    target_sheet.merge_cells(start_row = 1, start_column = cell_1[0], end_row = 1, end_column = target_sheet.max_column);

    # Tambah judul header
    target_sheet.cell(row = 1, column = cell_1[0]).value = "Stemmed";  



    # Update Workbook hasil (file xlsx)
    source_workbook.save(filename = FILEPATH_RESULT);

    # Buat heading setiap kolom menjadi bold
    set_heading_to_bold (source_workbook, target_sheet);

    # Return list token tanpa stopwords
    return token_stemmed_list;

"""
Modul untuk mendapatkan ringkasan data frekuensi kata
dari worksheet yang dipilih
"""
def get_frekuensi (source_workbook, source_sheet, target_sheet, cell_1, cell_2):
    # Mengambil source sheet
    source_sheet = source_workbook[source_sheet];
    
    # Buat sheet baru
    target_sheet = source_workbook.create_sheet(title = target_sheet);

    # Ambil tabel dijadikan list
    source_list = get_list_from_table(source_sheet, cell_1, cell_2);

    # Mengambil kordinat rentang tabel
    cell_1 = get_idx_coordinate(cell_1);
    cell_2 = get_idx_coordinate(cell_2);


    # Chain semua list dalam list source_list
    source_list = list(chain(*source_list));

    # Ubah menjadi object frequency
    frequency_list = FreqDist(source_list);

    # Urutkan semua frequensi kata berdasarkan valuenya
    sorted(frequency_list.items(), key = lambda item: item[1]);


    # Append Header Sheet
    target_sheet.append(["No", "Kata", "Frekuensi"]);

    # Append key value pair dari dictionary frequensi kata
    i = 1;
    for key in frequency_list:
        target_sheet.append([i, key, frequency_list[key]]);
        i = i + 1;



    # Update Workbook hasil (file xlsx)
    source_workbook.save(filename = FILEPATH_RESULT);

    # Buat heading setiap kolom menjadi bold
    set_heading_to_bold (source_workbook, target_sheet);


"""
Visualisasi frekuensi data
"""
def visualisasi_frekuensi (hasil_token_list, hasil_token_no_stopwords_list, hasil_token_no_stopwords_stemmed_list):
    ### Bandingkan frekuensi ###
    # Rangkai List of Token menjadi satu list
    hasil_token_list = list(chain(*hasil_token_list));
    hasil_token_no_stopwords_list = list(chain(*hasil_token_no_stopwords_list));
    hasil_token_no_stopwords_stemmed_list = list(chain(*hasil_token_no_stopwords_stemmed_list));


    # Buat tuple frekuensi
    freq_token_list = FreqDist(hasil_token_list);
    freq_token_no_stopwords_list = FreqDist(hasil_token_no_stopwords_list);
    freq_token_no_stopwords_stemmed_list = FreqDist(hasil_token_no_stopwords_stemmed_list);

    # Tampilkan 20 data paling sering muncul
    print("Tanpa Stopwords Removal:");
    print(np.array(freq_token_list.most_common(20)));
    print("\n\nDengan Stopwords Removal:");
    print(np.array(freq_token_no_stopwords_list.most_common(20)));
    print("\n\nDengan Stopwords Removal & Stemming:");
    print(np.array(freq_token_no_stopwords_stemmed_list.most_common(20)));



    # Visualisasi
    freq_token_list.plot(20, title = "Frekuensi Kata\nTanpa Stopwords Removal dan Stemming");
    freq_token_no_stopwords_list.plot(20, title = "Frekuensi Kata\ndengan Stopwords Removal");
    freq_token_no_stopwords_stemmed_list.plot(20, title = "Frekuensi Kata\ndengan Stopwords Removal dan Stemming");

    plt.show();
    




"""
MAIN FUNCTION
blok untuk memicu modul
"""
def main ():
    # Check file yang dibutuhkan
    if not path.exists(FILEPATH_RAW):
        if not path.exists("./DatasetText.xlsx"):
            print("File DatasetText.xlsx tidak ditemukan !\nPindahkan file DatasetText.xlsx ke dalam folder DATA");
            quit();
        else:
            makedirs("./DATA RESULT");
            copyfile("./DatasetText.xlsx", "./DATA RESULT/DatasetText.xlsx")
    

    # Ambil file DatasetText.xlsx (raw file)
    dataset = load_workbook(filename = "./DATA RESULT/DatasetText.xlsx", data_only=True);

    # Ambil sheetnya
    sheet_raw_dataset = dataset["Dataset"];

    # copy Dataset agar tidak merusak raw data
    # workbook_hasil merupakan file xlsx hasil pengolaan data
    workbook_hasil = copy_excel(sheet_raw_dataset, 201, FILEPATH_RESULT);

    # Tokenisasi 
    hasil_token_list = get_tokenize(workbook_hasil, "Dataset", "Tokenize", "B2", "B201");

    # Frekuensi kata dan Overall
    get_frekuensi(workbook_hasil, "Tokenize", "Freq Token", "B2", "AB201");

    # Case Folding
    set_table_to_lower_case(workbook_hasil, "Tokenize", "B2", "AB201");

    # Stopwords removal
    hasil_token_no_stopwords_list = get_stopwords_removal(workbook_hasil, "Tokenize", "No Stopwords", "B2", "AB201");

    # Frekuensi kata dan Overall
    get_frekuensi(workbook_hasil, "No Stopwords", "Freq no Stopwords", "B2", "AB201");

    # Stemming
    hasil_token_no_stopwords_stemmed_list = get_stemming(workbook_hasil, "No Stopwords", "Stemmed", "B2", "AB201");

    # Frekuensi kata dan Overall
    get_frekuensi(workbook_hasil, "Stemmed", "Final Result", "B2", "N201");


    # Visualisasi
    visualisasi_frekuensi(hasil_token_list, hasil_token_no_stopwords_list, hasil_token_no_stopwords_stemmed_list);

    # LIHAT DATA LENGKAP
    print("\nHasil Proses diatas dapat dilihat pada file : \"" + FILEPATH_RESULT + "\" !\n\n");




"""
Pemicu MAIN FUNCTION
"""
if __name__ == "__main__":
    main();











